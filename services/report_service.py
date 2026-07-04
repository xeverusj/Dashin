"""
services/report_service.py — Dashin Research Platform
Weekly campaign report generation.
Matches the exact format of the client's Excel report.
Outputs: in-portal view + downloadable XLSX.
"""

import io
from datetime import datetime, date, timedelta
from core.db import get_connection

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


# ── WEEK HELPERS ──────────────────────────────────────────────────────────────

def current_week_start() -> str:
    """Monday of current week as ISO string."""
    today = date.today()
    monday = today - timedelta(days=today.weekday())
    return monday.isoformat()


def week_label(week_start: str) -> str:
    """e.g. '12-19 Jan' from '2026-01-12'"""
    start = date.fromisoformat(week_start)
    end   = start + timedelta(days=6)
    if start.month == end.month:
        return f"{start.day}-{end.day} {start.strftime('%b')}"
    return f"{start.day} {start.strftime('%b')} - {end.day} {end.strftime('%b')}"


def get_week_range(week_start: str) -> tuple:
    start = date.fromisoformat(week_start)
    end   = start + timedelta(days=6)
    return start.isoformat(), end.isoformat()


# ── SAVE WEEKLY STATS ─────────────────────────────────────────────────────────

def save_weekly_stats(
    campaign_id:      int,
    week_start:       str,
    cold_emails_sent: int,
    followups_sent:   int,
    opens:            int,
    responded:        int,
    interested:       int,
    scheduled:        int,
    meetings_done:    int,
    entered_by:       int,
) -> int:
    """Save or update weekly stats for a campaign."""
    conn = get_connection()
    now  = datetime.utcnow().isoformat()
    _, week_end = get_week_range(week_start)
    label       = week_label(week_start)
    total_sent  = cold_emails_sent + followups_sent
    open_rate   = round(opens / total_sent * 100, 1) if total_sent > 0 else 0

    cur = conn.execute("""
        INSERT INTO campaign_weekly_stats
            (campaign_id, week_label, week_start, week_end,
             cold_emails_sent, followups_sent, total_sent,
             opens, open_rate, responded, interested,
             scheduled, meetings_done, entered_by, entered_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(campaign_id, week_label) DO UPDATE SET
            cold_emails_sent = excluded.cold_emails_sent,
            followups_sent   = excluded.followups_sent,
            total_sent       = excluded.total_sent,
            opens            = excluded.opens,
            open_rate        = excluded.open_rate,
            responded        = excluded.responded,
            interested       = excluded.interested,
            scheduled        = excluded.scheduled,
            meetings_done    = excluded.meetings_done,
            entered_by       = excluded.entered_by,
            entered_at       = excluded.entered_at
    """, (campaign_id, label, week_start, week_end,
          cold_emails_sent, followups_sent, total_sent,
          opens, open_rate, responded, interested,
          scheduled, meetings_done, entered_by, now))

    row_id = cur.lastrowid
    conn.commit()
    conn.close()
    return row_id


def get_weekly_stats(campaign_id: int) -> list:
    """All weekly stats for a campaign, oldest first."""
    conn = get_connection()
    rows = conn.execute("""
        SELECT * FROM campaign_weekly_stats
        WHERE campaign_id=?
        ORDER BY week_start ASC
    """, (campaign_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_campaign_totals(campaign_id: int) -> dict:
    """Aggregate totals across all weeks."""
    conn = get_connection()
    row = conn.execute("""
        SELECT
            SUM(cold_emails_sent) AS total_cold,
            SUM(followups_sent)   AS total_followups,
            SUM(total_sent)       AS total_sent,
            SUM(opens)            AS total_opens,
            SUM(responded)        AS total_responded,
            SUM(interested)       AS total_interested,
            SUM(scheduled)        AS total_scheduled,
            SUM(meetings_done)    AS total_meetings,
            COUNT(*)              AS weeks_active
        FROM campaign_weekly_stats WHERE campaign_id=?
    """, (campaign_id,)).fetchone()
    conn.close()

    if not row or not row["total_sent"]:
        return {}

    total_sent = row["total_sent"] or 0
    return {
        "total_cold":       row["total_cold"]      or 0,
        "total_followups":  row["total_followups"] or 0,
        "total_sent":       total_sent,
        "total_opens":      row["total_opens"]     or 0,
        "total_responded":  row["total_responded"] or 0,
        "total_interested": row["total_interested"]or 0,
        "total_scheduled":  row["total_scheduled"] or 0,
        "total_meetings":   row["total_meetings"]  or 0,
        "weeks_active":     row["weeks_active"]    or 0,
        "overall_open_rate":    round((row["total_opens"]    or 0) / total_sent * 100, 1),
        "overall_response_rate":round((row["total_responded"]or 0) / total_sent * 100, 1),
        "overall_meeting_rate": round((row["total_meetings"] or 0) / total_sent * 100, 2),
    }


# ── CRM SNAPSHOT ─────────────────────────────────────────────────────────────

def get_crm_snapshot(campaign_id: int) -> list:
    """
    Returns CRM data matching the original Excel CRM sheet format.
    {name, company, role, email, status, next_step, outreach_from,
     meeting_date, notes, last_updated}
    """
    conn = get_connection()
    rows = conn.execute("""
        SELECT
            l.full_name, l.title AS role,
            co.name     AS company,
            e.email,
            cl.crm_status   AS status,
            cl.next_step,
            cl.outreach_from,
            cl.meeting_date,
            cl.notes,
            cl.last_updated_at,
            u.name      AS last_updated_by_name
        FROM campaign_leads cl
        JOIN leads l      ON l.id  = cl.lead_id
        LEFT JOIN companies co ON co.id = l.company_id
        LEFT JOIN enrichment e ON e.lead_id = l.id
        LEFT JOIN users u ON u.id = cl.last_updated_by
        WHERE cl.campaign_id=?
        ORDER BY
            CASE cl.crm_status
                WHEN 'booked'            THEN 0
                WHEN 'meeting_requested' THEN 1
                WHEN 'interested'        THEN 2
                WHEN 'responded'         THEN 3
                WHEN 'waiting'           THEN 4
                WHEN 'contacted'         THEN 5
                ELSE 6
            END,
            l.full_name ASC
    """, (campaign_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── XLSX GENERATION ───────────────────────────────────────────────────────────

def generate_xlsx(campaign_id: int, campaign_name: str,
                  client_name: str) -> bytes | None:
    """
    Generate an XLSX report matching the client's Excel format.
    Sheet 1: CRM (like the CRM tab)
    Sheet 2+: One per week (like 12-19 Jan, 19-25 Jan etc.)
    Returns bytes or None if openpyxl not installed.
    """
    if not XLSX_AVAILABLE:
        return None

    wb     = openpyxl.Workbook()
    weekly = get_weekly_stats(campaign_id)
    crm    = get_crm_snapshot(campaign_id)
    totals = get_campaign_totals(campaign_id)

    # ── Styles ────────────────────────────────────────────────────────
    header_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1A1917")  # dark
    gold_fill      = PatternFill("solid", fgColor="C9A96E")  # gold
    alt_fill       = PatternFill("solid", fgColor="F8F7F4")  # cream
    border_side    = Side(style="thin", color="E0DDD8")
    thin_border    = Border(left=border_side, right=border_side,
                            bottom=border_side, top=border_side)
    center_align   = Alignment(horizontal="center", vertical="center")
    wrap_align     = Alignment(wrap_text=True, vertical="top")
    bold_font      = Font(name="Calibri", bold=True, size=10)
    normal_font    = Font(name="Calibri", size=10)

    def style_header(cell, text, fill=None):
        cell.value     = text
        cell.font      = header_font
        cell.fill      = fill or header_fill
        cell.alignment = center_align
        cell.border    = thin_border

    def style_cell(cell, value, bold=False, fill=None, align=None):
        cell.value     = value
        cell.font      = bold_font if bold else normal_font
        cell.border    = thin_border
        cell.alignment = align or Alignment(vertical="center")
        if fill:
            cell.fill = fill

    # ── CRM Sheet ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "CRM"

    # Title row
    ws.merge_cells("A1:K1")
    title_cell       = ws["A1"]
    title_cell.value = f"{client_name} — {campaign_name}"
    title_cell.font  = Font(name="Calibri", bold=True, size=14, color="1A1917")
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 30

    # Sub-header
    ws.merge_cells("A2:K2")
    sub              = ws["A2"]
    sub.value        = f"Generated: {datetime.now().strftime('%d %b %Y')} · Dashin Research"
    sub.font         = Font(name="Calibri", size=9, color="999999", italic=True)
    sub.alignment    = center_align

    # Headers
    crm_headers = ["Campaign", "Status", "Name", "Company", "Role",
                   "Email", "Next Step", "Outreach From",
                   "Meeting Date", "Comment", "Last Updated"]
    for col, h in enumerate(crm_headers, 1):
        style_header(ws.cell(3, col), h)
    ws.row_dimensions[3].height = 20

    # Status colour map
    status_colors = {
        "booked":            "C6EFCE",
        "meeting_requested": "FFEB9C",
        "interested":        "BDD7EE",
        "responded":         "E2EFDA",
        "waiting":           "FFF2CC",
        "contacted":         "F2F2F2",
        "not_interested":    "FCE4D6",
        "no_show":           "F4CCCC",
    }

    for i, row in enumerate(crm, 1):
        r    = i + 3
        fill = PatternFill("solid", fgColor=status_colors.get(
            row.get("status",""), "FFFFFF"))
        style_cell(ws.cell(r, 1),  campaign_name, fill=fill)
        style_cell(ws.cell(r, 2),  (row.get("status") or "").replace("_"," ").title(), fill=fill)
        style_cell(ws.cell(r, 3),  row.get("full_name",""), bold=True, fill=fill)
        style_cell(ws.cell(r, 4),  row.get("company",""),   fill=fill)
        style_cell(ws.cell(r, 5),  row.get("role",""),      fill=fill)
        style_cell(ws.cell(r, 6),  row.get("email",""),     fill=fill)
        style_cell(ws.cell(r, 7),  row.get("next_step",""), fill=fill)
        style_cell(ws.cell(r, 8),  row.get("outreach_from",""), fill=fill)
        style_cell(ws.cell(r, 9),  row.get("meeting_date",""),  fill=fill)
        style_cell(ws.cell(r, 10), row.get("notes",""),
                   fill=fill, align=wrap_align)
        style_cell(ws.cell(r, 11),
                   (row.get("last_updated_at","") or "")[:10], fill=fill)
        ws.row_dimensions[r].height = 18

    # Column widths
    col_widths = [22, 20, 22, 22, 22, 28, 22, 24, 14, 40, 14]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"

    # ── Weekly Stats Sheets ────────────────────────────────────────────
    for week in weekly:
        ws2       = wb.create_sheet(title=week["week_label"][:31])
        label     = week["week_label"]
        camp_name_cell = campaign_name[:20]

        # Header block (matches original format)
        ws2.merge_cells("A1:B1")
        ws2["A1"].value = "Sent from email"
        ws2["A1"].font  = bold_font

        ws2.merge_cells("C1:I1")
        ws2["C1"].value     = camp_name_cell
        ws2["C1"].font      = Font(name="Calibri", bold=True, size=11)
        ws2["C1"].alignment = center_align

        # Column headers
        stat_headers = ["", "Campaign", "Email Sent (Cold)",
                        "Open", "Open Rate (%)", "Responded",
                        "Interested (Pipeline)", "Scheduled", "Meetings Done"]
        for col, h in enumerate(stat_headers, 1):
            c      = ws2.cell(2, col)
            c.value = h
            c.font  = bold_font
            if col > 1:
                c.fill      = gold_fill if col <= 5 else header_fill
                c.font      = Font(name="Calibri", bold=True,
                                   color="FFFFFF" if col > 5 else "1A1917",
                                   size=10)
                c.alignment = center_align
                c.border    = thin_border

        # Data rows
        rows_data = [
            ("Cold Email", week["cold_emails_sent"], "",
             "", week["responded"], week["interested"],
             "", ""),
            ("Follow ups", week["followups_sent"],    "", "", "", "", "", ""),
            ("TOTAL:",     week["total_sent"],
             week["opens"], f"{week['open_rate']}%",
             week["responded"], week["interested"],
             week["scheduled"], week["meetings_done"]),
        ]

        for i, (label_text, *vals) in enumerate(rows_data, 3):
            ws2.cell(i, 1).value = ""
            ws2.cell(i, 2).value = label_text
            ws2.cell(i, 2).font  = bold_font if label_text == "TOTAL:" else normal_font
            for col, val in enumerate(vals, 3):
                c       = ws2.cell(i, col)
                c.value = val
                c.font  = bold_font if label_text == "TOTAL:" else normal_font
                c.alignment = center_align
                c.border    = thin_border
                if label_text == "TOTAL:":
                    c.fill = PatternFill("solid", fgColor="E8F5E9")

        # Column widths
        for col, w in enumerate([18, 22, 18, 10, 14, 12, 20, 12, 14], 1):
            ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Summary Sheet ──────────────────────────────────────────────────
    if totals:
        ws3       = wb.create_sheet(title="Summary")
        ws3.merge_cells("A1:C1")
        ws3["A1"].value     = f"Campaign Summary — {campaign_name}"
        ws3["A1"].font      = Font(name="Calibri", bold=True, size=13)
        ws3["A1"].alignment = center_align

        summary_rows = [
            ("Total Emails Sent",    totals.get("total_sent",0)),
            ("Cold Emails",          totals.get("total_cold",0)),
            ("Follow Ups",           totals.get("total_followups",0)),
            ("Total Opens",          totals.get("total_opens",0)),
            ("Open Rate",            f"{totals.get('overall_open_rate',0)}%"),
            ("Total Responded",      totals.get("total_responded",0)),
            ("Interested / Pipeline",totals.get("total_interested",0)),
            ("Meetings Scheduled",   totals.get("total_scheduled",0)),
            ("Meetings Done",        totals.get("total_meetings",0)),
            ("Response Rate",        f"{totals.get('overall_response_rate',0)}%"),
            ("Weeks Active",         totals.get("weeks_active",0)),
        ]

        for i, (label_text, val) in enumerate(summary_rows, 3):
            ws3.cell(i, 1).value = label_text
            ws3.cell(i, 1).font  = bold_font
            ws3.cell(i, 2).value = val
            ws3.cell(i, 2).font  = normal_font
            ws3.cell(i, 2).alignment = center_align
            if i % 2 == 0:
                ws3.cell(i, 1).fill = alt_fill
                ws3.cell(i, 2).fill = alt_fill

        ws3.column_dimensions["A"].width = 28
        ws3.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── REPORT SUMMARY FOR PORTAL ─────────────────────────────────────────────────

def get_campaign_report_data(campaign_id: int) -> dict:
    """
    Full report data for displaying in client portal or campaign manager view.
    """
    conn = get_connection()

    campaign = conn.execute("""
        SELECT ca.*, cl.name AS client_name
        FROM campaigns ca
        LEFT JOIN clients cl ON cl.id = ca.client_id
        WHERE ca.id=?
    """, (campaign_id,)).fetchone()

    conn.close()
    if not campaign:
        return {}

    return {
        "campaign":     dict(campaign),
        "weekly_stats": get_weekly_stats(campaign_id),
        "totals":       get_campaign_totals(campaign_id),
        "crm":          get_crm_snapshot(campaign_id),
    }
